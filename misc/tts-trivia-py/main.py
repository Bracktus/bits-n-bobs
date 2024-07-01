from random import shuffle
from textwrap import TextWrapper

import cairo
from openpyxl import load_workbook

from font_loader import create_cairo_font_face_for_file

CATEGORIES = ["geography", "entertainment", "history", "art", "science", "sport"]
TAU = 6.2831
WIDTH = int(720 * 1.5)
HEIGHT = int(500 * 1.5)

norm = lambda col: [v / 255 for v in col]  # noqa: E731
BLUE = norm([53, 79, 157])
PINK = norm([234, 66, 161])
YELLOW = norm([234, 226, 66])
BROWN = norm([64, 34, 34])
GREEN = norm([21, 69, 5])
ORANGE = norm([210, 93, 3])


def read_qas(filename="Trivia.xlsx"):
    wb = load_workbook(filename)
    ws = wb.active

    if ws is None:
        raise Exception("Worksheet is empty")

    rows = list(
        row for row in ws.iter_rows(min_row=2, max_col=6, values_only=True) if all(row)
    )

    questions = [
        map(str, question)
        for idx, question in enumerate(rows)
        if idx % 2 == 0  # noqa: E501
    ]
    answers = [
        map(str, answer)
        for idx, answer in enumerate(rows)
        if idx % 2 == 1  # noqa: E501
    ]

    final = {cat: list() for cat in CATEGORIES}
    mapping = {i: v for i, v in enumerate(CATEGORIES)}

    for que, ans in zip(questions, answers):
        for i, (q, a) in enumerate(zip(que, ans)):
            card = {}
            card["question"] = q
            card["answer"] = a
            category = mapping[i]
            final[category].append(card)

    return final


def draw_card(card, type, card_name):
    """A card is a set of 5 pieces of text
    It can be an answer card or a question card"""
    surface = cairo.ImageSurface(cairo.FORMAT_ARGB32, WIDTH, HEIGHT)
    ctx = cairo.Context(surface)

    # Setting bg colour
    ctx.set_source_rgb(1, 1, 1)  # White
    ctx.rectangle(0, 0, WIDTH, HEIGHT)
    ctx.fill()

    # font settings
    face = create_cairo_font_face_for_file(
        "./font/FrankRuhlLibre-VariableFont_wght.ttf", 0
    )
    ctx.set_font_face(face)
    ctx.set_font_size(25)

    def ellipse(x, y, col):
        ctx.arc(x, y, HEIGHT * 0.05, 0, 6.28)
        ctx.set_source_rgb(*col)
        ctx.fill()

    X_MARGIN = WIDTH * 0.075
    Y_MARGIN = HEIGHT * 0.2

    cols = [BLUE, PINK, YELLOW, BROWN, GREEN, ORANGE]
    ctx.translate(0, HEIGHT * 0.1)

    wrapper = TextWrapper()
    wrapper.width = 75

    for idx, (text, col) in enumerate(zip(card, cols)):
        y = ((HEIGHT - Y_MARGIN) / 5) * idx

        ellipse(X_MARGIN, y, col)
        ctx.move_to(X_MARGIN + WIDTH * 0.05, y)
        ctx.set_source_rgb(0, 0, 0)  # Black

        wrapped_text = wrapper.wrap(text)
        for wrap_idx, line in enumerate(wrapped_text):
            ctx.move_to(X_MARGIN + WIDTH * 0.05, y + wrap_idx * 30)
            ctx.text_path(line)
            ctx.fill()

    if type == "answer":
        surface.write_to_png(f"cards/answers/{card_name}")
    else:
        surface.write_to_png(f"cards/questions/{card_name}")


def gen_qas():
    questions = read_qas()
    for key in questions:
        shuffle(questions[key])

    # While there are still questions
    while questions["geography"]:
        q_card = []
        a_card = []
        for cat in CATEGORIES:
            qa = questions[cat].pop()
            q_card.append(qa["question"])
            a_card.append(qa["answer"])

        yield q_card, a_card


for idx, (q_card, a_card) in enumerate(gen_qas()):
    draw_card(q_card, "question", f"{idx}_question.png")
    draw_card(a_card, "answer", f"{idx}_answer.png")
